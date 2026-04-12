"""
数据导出路由 — 库存、销售、批次回本数据导出为 Excel 文件。

所有导出端点返回 .xlsx 文件（StreamingResponse），使用 openpyxl 在内存中生成。
"""

import datetime
import io
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, func
from sqlalchemy.orm import Session, selectinload

from database import get_db
from models import Batch, DictMaterial, Item, SaleRecord
from schemas import DictTagOut, ItemImageOut, ItemSpecOut

router = APIRouter(prefix="/export", tags=["数据导出"])

# ──────────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────────

def _today_str() -> str:
    return datetime.date.today().strftime("%Y%m%d")


def _write_headers(ws, headers, row=1):
    """写入表头行（加粗）"""
    from openpyxl.styles import Font
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True)


def _write_row(ws, row, values):
    """写入一行数据"""
    for col_idx, v in enumerate(values, 1):
        ws.cell(row=row, column=col_idx, value=v)


def _load_item_options():
    return [
        selectinload(Item.material),
        selectinload(Item.type),
        selectinload(Item.batch),
        selectinload(Item.supplier),
        selectinload(Item.tags),
        selectinload(Item.images),
        selectinload(Item.spec),
    ]


_STATUS_MAP = {"in_stock": "在库", "sold": "已售", "returned": "已退", "lent_out": "借出"}
_CHANNEL_MAP = {"store": "门店", "wechat": "微信", "ecommerce": "电商"}
_BATCH_STATUS_MAP = {"new": "未开始", "selling": "销售中", "paid_back": "已回本", "cleared": "清仓完毕"}


# ──────────────────────────────────────────────
# 1. 库存导出
# ──────────────────────────────────────────────

@router.get(
    "/inventory",
    summary="导出库存数据",
    description="导出库存货品列表为 Excel 文件。",
)
def export_inventory(
    material_id: Optional[int] = Query(None, description="按材质 ID 筛选"),
    type_id: Optional[int] = Query(None, description="按器型 ID 筛选"),
    status: Optional[str] = Query(None, description="按状态筛选"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    supplier_id: Optional[int] = Query(None, description="按供货商 ID 筛选"),
    counter: Optional[int] = Query(None, description="按柜台号筛选"),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    import openpyxl

    filters = [Item.is_deleted == False]
    if material_id is not None:
        filters.append(Item.material_id == material_id)
    if type_id is not None:
        filters.append(Item.type_id == type_id)
    if status is not None:
        filters.append(Item.status == status)
    if supplier_id is not None:
        filters.append(Item.supplier_id == supplier_id)
    if counter is not None:
        filters.append(Item.counter == counter)
    if keyword:
        kw = f"%{keyword}%"
        filters.append(or_(
            Item.sku_code.ilike(kw), Item.name.ilike(kw),
            Item.batch_code.ilike(kw), Item.cert_no.ilike(kw), Item.notes.ilike(kw),
        ))

    items = (
        db.query(Item)
        .options(*_load_item_options())
        .filter(*filters)
        .order_by(Item.created_at.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存数据"

    headers = ["SKU编号", "名称", "批次号", "材质", "器型", "成本/分摊成本", "底价",
               "零售价", "状态", "柜台", "进货日期", "在库天数", "备注"]
    _write_headers(ws, headers)

    today = datetime.date.today()
    for idx, item in enumerate(items, 2):
        cost = item.allocated_cost if item.allocated_cost is not None else item.cost_price
        age = None
        ref_date = item.purchase_date or item.created_at.date() if item.created_at else None
        if ref_date:
            age = (today - ref_date).days
        tag_names = ", ".join(t.name for t in item.tags) if item.tags else ""
        notes = item.notes or ""
        if tag_names:
            notes = f"{tag_names} | {notes}" if notes else tag_names

        _write_row(ws, idx, [
            item.sku_code,
            item.name or "",
            item.batch_code or "",
            item.material.name,
            item.type.name if item.type else "",
            round(cost, 2) if cost else 0,
            round(item.floor_price, 2) if item.floor_price else "",
            round(item.selling_price, 2) if item.selling_price else 0,
            _STATUS_MAP.get(item.status, item.status),
            item.counter or "",
            str(item.purchase_date) if item.purchase_date else "",
            age or "",
            notes,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"库存导出_{_today_str()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ──────────────────────────────────────────────
# 2. 销售导出
# ──────────────────────────────────────────────

@router.get(
    "/sales",
    summary="导出销售数据",
    description="导出销售记录为 Excel 文件。",
)
def export_sales(
    start_date: Optional[datetime.date] = Query(None, description="成交日期起（含）"),
    end_date: Optional[datetime.date] = Query(None, description="成交日期止（含）"),
    channel: Optional[str] = Query(None, description="渠道筛选"),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    import openpyxl

    filters = []
    if start_date is not None:
        filters.append(SaleRecord.sale_date >= start_date)
    if end_date is not None:
        filters.append(SaleRecord.sale_date <= end_date)
    if channel is not None:
        filters.append(SaleRecord.channel == channel)

    records = (
        db.query(SaleRecord)
        .options(
            selectinload(SaleRecord.item).selectinload(Item.material),
            selectinload(SaleRecord.customer),
        )
        .filter(*filters)
        .order_by(SaleRecord.sale_date.desc(), SaleRecord.id.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "销售数据"

    headers = ["销售单号", "SKU编号", "货品名称", "材质", "成交价", "成本", "毛利",
               "渠道", "销售日期", "客户", "备注"]
    _write_headers(ws, headers)

    for idx, r in enumerate(records, 2):
        item = r.item
        cost = item.allocated_cost if item.allocated_cost is not None else item.cost_price
        if cost is None:
            cost = 0.0
        gross_profit = r.actual_price - cost

        _write_row(ws, idx, [
            r.sale_no,
            item.sku_code,
            item.name or "",
            item.material.name,
            round(r.actual_price, 2),
            round(cost, 2),
            round(gross_profit, 2),
            _CHANNEL_MAP.get(r.channel, r.channel),
            str(r.sale_date),
            r.customer.name if r.customer else "",
            r.note or "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"销售导出_{_today_str()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ──────────────────────────────────────────────
# 3. 批次回本导出
# ──────────────────────────────────────────────

@router.get(
    "/batches",
    summary="导出批次回本数据",
    description="导出批次回本状态为 Excel 文件。",
)
def export_batches(
    material_id: Optional[int] = Query(None, description="按材质ID筛选"),
    status: Optional[str] = Query(None, description="按批次状态筛选"),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    import openpyxl

    query = db.query(Batch)
    if material_id is not None:
        query = query.filter(Batch.material_id == material_id)

    batches = query.order_by(Batch.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批次回本"

    headers = ["批次编号", "材质", "总成本", "数量", "已售数", "已回款", "利润",
               "回本进度%", "状态"]
    _write_headers(ws, headers)

    row_idx = 2
    for batch in batches:
        # 复用 dashboard.py 中的批次统计逻辑
        items_query = db.query(Item).filter(
            Item.batch_id == batch.id, Item.is_deleted == False
        )
        items_count = items_query.count()

        sold_subquery = (
            db.query(SaleRecord.item_id)
            .join(Item, SaleRecord.item_id == Item.id)
            .filter(Item.batch_id == batch.id, Item.is_deleted == False)
            .subquery()
        )
        sold_count = db.query(func.count()).select_from(sold_subquery).scalar() or 0

        revenue_result = (
            db.query(func.sum(SaleRecord.actual_price))
            .join(Item, SaleRecord.item_id == Item.id)
            .filter(Item.batch_id == batch.id, Item.is_deleted == False)
            .scalar()
        )
        revenue = float(revenue_result) if revenue_result else 0.0

        profit = revenue - batch.total_cost
        payback_rate = revenue / batch.total_cost if batch.total_cost > 0 else 0.0

        if sold_count == 0:
            batch_status = "new"
        elif payback_rate >= 1.0:
            batch_status = "cleared" if sold_count >= batch.quantity else "paid_back"
        else:
            batch_status = "selling"

        # 按状态筛选
        if status is not None and batch_status != status:
            continue

        _write_row(ws, row_idx, [
            batch.batch_code,
            batch.material.name,
            round(batch.total_cost, 2),
            batch.quantity,
            sold_count,
            round(revenue, 2),
            round(profit, 2),
            round(payback_rate * 100, 1),
            _BATCH_STATUS_MAP.get(batch_status, batch_status),
        ])
        row_idx += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"批次回本_{_today_str()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
